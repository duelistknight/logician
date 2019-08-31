import pandas as pd 
import tkinter as tk
import openpyxl as op

a=tk.Tk()
a.title("E&Y Form")
file = r'city.xls'
df = pd.read_excel(file)

city1,state1,nation=[0,0,0]

#------------------ENTRY---------------------------
country_label=tk.Label(text='Country Name')
country_label.grid(row=0,column=0)

var_country=tk.StringVar(a)
var_country.set('India')
country=tk.Entry(a, textvariable=var_country)
country.grid(row=0, column=1)
     
#---------------------Comment------------------------------------
comment_label=tk.Label(text='Comment')
comment_label.grid(row=4,column=0)
var_comment=tk.StringVar(a)
comment=tk.Entry(a, textvariable=var_comment)
comment.grid(row=4, column=1)

#_------------------states and city-----------------------------------------------------------

dicti = {}
states= []
for i in df['State']:
        states.append(i)
states=set(states)
states=tuple(states)

dist=[]
        
for j in range(0,(len(states))):
        af = df[df['State'] == states[j]]
        for i in af['Name of City']:
                dist.append(i)
        s={states[j]:dist}
        dicti.update(s)
        dist=[]

variable_a = tk.StringVar()
variable_b = tk.StringVar()

def update_options(*args):
        a = dicti[variable_a.get()]
        variable_b.set(a[0])
        state1=variable_a.get()
        menu = optionmenu_b['menu']
        menu.delete(0, 'end')

        for country in a:
            menu.add_command(label=country, command=lambda nation=country: variable_b.set(nation))

variable_a.trace('w', update_options)

optionmenu_a = tk.OptionMenu(a, variable_a, *dicti.keys())
optionmenu_b = tk.OptionMenu(a, variable_b, '')
variable_a.set('Karnataka')
state_label=tk.Label(text='Select State')
state_label.grid(row=1,column=0)
city_label=tk.Label(text='Select City')
city_label.grid(row=2,column=0)

optionmenu_a.grid(row=1,column=1)
optionmenu_b.grid(row=2,column=1)
#-----------------------------------Poputation---------------------------
n_city=[]
popu=[]
for i in df['Name of City']:
        n_city.append(i)
for i in df['Population (2011)']:
        popu.append(i)

var_pop=tk.StringVar(a)
var_pop.set('0')
populatio=0
def new(*args):
        populatio=popu[n_city.index(variable_b.get())]
        populatio=str(int(populatio))
        var_pop.set(populatio)
        city=variable_b.get()
        print(populatio)
         
variable_b.trace('w', new)
population=tk.Entry(a, textvariable=var_pop)
def diploy():
        city1=str(variable_b.get())
        state1=str(variable_a.get())
        rows=(var_country.get(),state1,city1,population.get(),comment.get())
        try:
                wb = op.load_workbook('E&Y data.xlsx')
                ws = wb.get_sheet_by_name('Sheet')
                ws.append(rows)
                wb.save('E&Y data.xlsx')
                wb.close()
        except Exception:
                book=op.Workbook()
                sheet=book.active
                row=('Country','State', 'City','Population','Comment')
                sheet.append(row)
                book.save('E&Y data.xlsx')
                wb = op.load_workbook('E&Y data.xlsx')
                ws = wb.get_sheet_by_name('Sheet')
                ws.append(rows)
                wb.save('E&Y data.xlsx')
                wb.close()
        var_country.set('India')
        var_pop.set('0')
        var_comment.set('')
        variable_a.set('Karnataka')
        variable_b.set('Bengaluru')
        
                
country_label=tk.Label(text='Population')
country_label.grid(row=3,column=0)
population.grid(row=3, column=1)

c=tk.Button(a, text= 'SUBMIT', command=diploy)
c.grid(row=5,column=2)

a.mainloop()
